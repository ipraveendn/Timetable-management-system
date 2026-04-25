import random
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import JSONResponse, StreamingResponse
from database import supabase
from dependencies import get_college_id, get_current_user
from io import BytesIO
from openpyxl import Workbook
from conflict_validator import ConflictValidator
from datetime import datetime, timedelta

router = APIRouter(tags=["Timetable Engine"])

def calculate_slots(start_time_str, slots_count, duration_mins):
    """Generate dynamic time slots based on college config."""
    try:
        start_time = datetime.strptime(start_time_str, "%H:%M")
        slots = []
        for i in range(slots_count):
            slot_start = start_time + timedelta(minutes=i * duration_mins)
            slots.append(slot_start.strftime("%H:%M"))
        return slots
    except:
        return ["09:00", "10:00", "11:00", "12:00", "13:00", "14:00", "15:00", "16:00"]

@router.post("/generate-timetable")
async def generate_timetable(
    college_id: str = Depends(get_college_id),
    current_user: dict = Depends(get_current_user)
):
    """
    Generate conflict-free timetable using logic and feature flags.
    Checks 5 locks and 6 rules for every assignment.
    """
    if current_user["role"] not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Only admin can generate timetable")

    try:
        # 1. Fetch data
        faculty_res = supabase.table("faculty").select("*").eq("college_id", college_id).eq("status", "active").execute()
        subjects_res = supabase.table("subjects").select("*").eq("college_id", college_id).execute()
        rooms_res = supabase.table("rooms").select("*").eq("college_id", college_id).execute()
        
        if not faculty_res.data or not subjects_res.data:
            raise HTTPException(status_code=400, detail="Faculty or Subjects data missing. Upload Excel first.")

        faculty = faculty_res.data
        subjects = subjects_res.data
        rooms = rooms_res.data or []
        
        # 2. Get College Flags
        from feature_flags import DEFAULT_FLAGS
        flags = dict(DEFAULT_FLAGS)
        try:
            flags_res = supabase.table("feature_flags").select("*").eq("college_id", college_id).execute()
            if flags_res.data:
                stored = flags_res.data[0]
                flags.update({k: stored[k] for k in DEFAULT_FLAGS if k in stored})
        except Exception:
            pass # Continue with defaults
        
        # 3. Setup Locks (In-Memory)
        faculty_time_lock = {f["id"]: {} for f in faculty}
        room_time_lock = {r["id"]: {} for r in rooms}
        faculty_load_lock = {f["id"]: {} for f in faculty}
        subject_count_lock = {s["id"]: 0 for s in subjects}
        subject_day_lock = {s["id"]: set() for s in subjects}
        
        # 4. Generate Days and Time Slots
        days_base = ["Mon", "Tue", "Wed", "Thu", "Fri"]
        if flags.get("saturday_enabled"): days_base.append("Sat")
        if flags.get("sunday_enabled"): days_base.append("Sun")
        
        time_slots_base = calculate_slots(
            flags.get("start_time", "09:00"),
            flags.get("slots_per_day", 8),
            flags.get("slot_duration_mins", 60)
        )
        
        if flags.get("break_after_3rd_period", True) and len(time_slots_base) > 3:
            # Usually the 4th slot (index 3) is a break
            time_slots_base.pop(3)
        
        global_max_lectures = flags.get("max_lectures_per_day", 4)
        
        # 5. Core Scheduling Algorithm (Backtrack-lite with Shuffling)
        generated_slots = []
        unassigned_classes = []
        
        # Sort subjects: prioritize those with high frequency
        sorted_subjects = sorted(subjects, key=lambda s: s.get("classes_per_week", 4), reverse=True)
        
        for subject in sorted_subjects:
            semester = subject.get("semester", 1)
            sub_id = subject["id"]
            classes_needed = subject.get("classes_per_week", 2)
            
            # Find matching faculty (Expertise + Semester)
            eligible_faculty = [f for f in faculty if 
                                semester in f.get("semesters", [1,2,3,4,5,6,7,8]) and 
                                (subject.get("name") in f.get("subjects", []) or not f.get("subjects"))]
            
            if not eligible_faculty:
                unassigned_classes.append(f"{subject.get('name')} (No eligible faculty)")
                continue

            for class_num in range(classes_needed):
                assigned = False
                
                # Shuffle days and times for randomness
                shuffled_days = list(days_base)
                random.shuffle(shuffled_days)
                
                for day in shuffled_days:
                    if assigned: break
                    if day in subject_day_lock[sub_id]: continue # Rule: Same subject only once a day
                    
                    shuffled_times = list(time_slots_base)
                    random.shuffle(shuffled_times)
                    
                    for time_slot in shuffled_times:
                        if assigned: break
                        
                        # Rule: Evenly distribute load
                        if flags.get("even_distribution", True):
                            faculty_list = sorted(eligible_faculty, key=lambda f: sum(faculty_load_lock[f["id"]].values()))
                        else:
                            faculty_list = list(eligible_faculty)
                            random.shuffle(faculty_list)
                            
                        for f in faculty_list:
                            fid = f["id"]
                            
                            # Lock 1: Faculty Time check
                            day_fac_slots = faculty_time_lock[fid].setdefault(day, set())
                            if time_slot in day_fac_slots: continue
                            
                            # Lock 2: Faculty Load check
                            day_load = faculty_load_lock[fid].setdefault(day, 0)
                            f_max = min(f.get("max_classes_per_day", 5), global_max_lectures)
                            if day_load >= f_max: continue
                            
                            # Lock 3: Room check (if rooms exist)
                            assigned_room_id = None
                            if rooms:
                                matching_rooms = [r for r in rooms if not subject.get("room_type_required") or r.get("room_type") == subject.get("room_type_required")]
                                random.shuffle(matching_rooms)
                                for r in matching_rooms:
                                    rid = r["id"]
                                    day_room_slots = room_time_lock[rid].setdefault(day, set())
                                    if time_slot not in day_room_slots:
                                        assigned_room_id = rid
                                        break
                                if not assigned_room_id: continue # No room available for this slot
                            
                            # ALL RULES PASSED: COMMIT ASSIGNMENT
                            faculty_time_lock[fid][day].add(time_slot)
                            faculty_load_lock[fid][day] += 1
                            if assigned_room_id: room_time_lock[assigned_room_id][day].add(time_slot)
                            subject_day_lock[sub_id].add(day)
                            
                            # End time calculation
                            st = datetime.strptime(time_slot, "%H:%M")
                            et = (st + timedelta(minutes=flags.get("slot_duration_mins", 60))).strftime("%H:%M")
                            
                            generated_slots.append({
                                "college_id": college_id,
                                "semester": semester,
                                "day": day,
                                "start_time": time_slot,
                                "end_time": et,
                                "subject_id": sub_id,
                                "faculty_id": fid,
                                "room_id": assigned_room_id,
                                "is_substituted": False
                            })
                            assigned = True
                            break
                            
                if not assigned:
                    unassigned_classes.append(f"{subject.get('name')} - Class #{class_num+1} (No free slot)")

        # 6. Final Validation
        conflicts = ConflictValidator.validate(generated_slots, faculty, subjects)
        if conflicts:
            return JSONResponse(status_code=400, content={
                "message": "Generated timetable has conflicts. Try again.",
                "conflicts": conflicts
            })

        # 7. Safe Database Update
        # Transaction-like approach: Clear current slots and insert new ones
        supabase.table("timetable_slots").delete().eq("college_id", college_id).execute()
        
        # Batch insert to avoid too many requests
        for i in range(0, len(generated_slots), 50):
            batch = generated_slots[i:i+50]
            supabase.table("timetable_slots").insert(batch).execute()

        return {
            "success": True,
            "message": f"Generated {len(generated_slots)} classes successfully.",
            "unassigned_count": len(unassigned_classes),
            "unassigned_details": unassigned_classes[:10] # Show first 10
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Generation crash: {str(e)}")

@router.get("/timetable")
async def get_timetable(semester: int, college_id: str = Depends(get_college_id)):
    try:
        res = supabase.table("timetable_slots").select("*").eq("college_id", college_id).eq("semester", semester).execute()
        return {"timetable": res.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching timetable: {str(e)}")

@router.get("/timetable/faculty/{faculty_id}")
async def get_faculty_timetable(faculty_id: str, college_id: str = Depends(get_college_id)):
    try:
        res = supabase.table("timetable_slots").select("*").eq("college_id", college_id).eq("faculty_id", faculty_id).execute()
        return {"timetable": res.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching faculty timetable: {str(e)}")

@router.get("/export/timetable")
async def export_timetable(semester: int, format: str = "excel", college_id: str = Depends(get_college_id)):
    try:
        res = supabase.table("timetable_slots").select("*").eq("college_id", college_id).eq("semester", semester).execute()
        timetable_data = res.data
        
        if format == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = f"Semester {semester} Timetable"
            
            headers = ["Day", "Time", "Subject", "Faculty", "Room"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            for row_idx, slot in enumerate(timetable_data, 2):
                ws.cell(row=row_idx, column=1, value=slot["day"])
                ws.cell(row=row_idx, column=2, value=slot["start_time"])
                
                subject_res = supabase.table("subjects").select("name").eq("id", slot["subject_id"]).execute()
                subject_name = subject_res.data[0]["name"] if subject_res.data else "Unknown"
                ws.cell(row=row_idx, column=3, value=subject_name)
                
                faculty_res = supabase.table("faculty").select("name").eq("id", slot["faculty_id"]).execute()
                faculty_name = faculty_res.data[0]["name"] if faculty_res.data else "Unknown"
                ws.cell(row=row_idx, column=4, value=faculty_name)
                
                room_res = supabase.table("rooms").select("room_code").eq("id", slot["room_id"]).execute()
                room_name = room_res.data[0]["room_code"] if room_res.data else "Unknown"
                ws.cell(row=row_idx, column=5, value=room_name)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"semester_{semester}_timetable.xlsx"
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        else:
            raise HTTPException(status_code=400, detail="Invalid format specified")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting timetable: {str(e)}")
