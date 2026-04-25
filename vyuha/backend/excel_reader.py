from fastapi import APIRouter, UploadFile, File, HTTPException, Request, Depends
from openpyxl import load_workbook
from database import supabase
from dependencies import get_college_id
import re

router = APIRouter()


def _split_csv(value):
    if not value:
        return []
    return [item.strip() for item in str(value).split(",") if item and str(item).strip()]


def _norm_day(day_value):
    day_map = {
        "monday": "Mon", "mon": "Mon",
        "tuesday": "Tue", "tue": "Tue", "tues": "Tue",
        "wednesday": "Wed", "wed": "Wed",
        "thursday": "Thu", "thu": "Thu", "thurs": "Thu",
        "friday": "Fri", "fri": "Fri",
        "saturday": "Sat", "sat": "Sat",
        "sunday": "Sun", "sun": "Sun",
    }
    raw = str(day_value).strip()
    if not raw:
        return None
    return day_map.get(raw.lower(), raw[:3].title())


def _norm_days(value):
    days = [_norm_day(item) for item in _split_csv(value)]
    return [d for d in days if d]


def _norm_room_type(value):
    raw = str(value).strip().lower()
    if not raw:
        return "classroom"
    if "lab" in raw:
        return "lab"
    if "hall" in raw:
        return "hall"
    if "work" in raw:
        return "workshop"
    return "classroom"


def _chunk_rows(rows, chunk_size=50):
    for start in range(0, len(rows), chunk_size):
        yield rows[start:start + chunk_size]


def _batch_insert(table_name, rows, chunk_size=50):
    if not rows:
        return
    for chunk in _chunk_rows(rows, chunk_size):
        supabase.table(table_name).insert(chunk).execute()

@router.post("/upload-excel")
async def upload_excel(
    request: Request,
    file: UploadFile = File(...),
    replace_existing: bool = True,
    college_id: str = Depends(get_college_id)
):
    try:
        if not supabase:
            raise HTTPException(status_code=500, detail="Database connection failed")
            
        wb = load_workbook(filename=file.file)
        
        email_cache = {}
        
        # Process Faculty
        faculty_sheet = wb["Faculty"] if "Faculty" in wb.sheetnames else None
        faculty_data = []
        if faculty_sheet:
            for row in faculty_sheet.iter_rows(min_row=2, values_only=True):
                if row[0]: # Name
                    email_val = str(row[7]) if len(row) > 7 and row[7] else ""
                    faculty_data.append({
                        "college_id": college_id,
                        "name": str(row[0]).strip()[:100],
                        "employee_id": str(row[1]).strip()[:50] if row[1] else "",
                        "subjects": _split_csv(row[2]),
                        "semesters": [int(x) for x in _split_csv(row[3])] if row[3] else [],
                        "max_classes_per_day": int(row[4]) if row[4] else 4,
                        "available_days": _norm_days(row[5]),
                        "department": str(row[6]).strip()[:100] if row[6] else "",
                        "email": email_val
                    })
                    
                    if email_val:
                        email_cache[row[0]] = email_val
        
        # Process Subjects
        subjects_sheet = wb["Subjects"] if "Subjects" in wb.sheetnames else None
        subjects_data = []
        if subjects_sheet:
            for row in subjects_sheet.iter_rows(min_row=2, values_only=True):
                if row[0]: # Name
                    subjects_data.append({
                        "college_id": college_id,
                        "semester": int(row[0]),
                        "name": str(row[1]).strip()[:100],
                        "classes_per_week": int(row[2]) if len(row)>2 and row[2] else 0,
                        "room_type_required": _norm_room_type(row[3]) if len(row) > 3 and row[3] else "classroom",
                        "duration_minutes": int(row[4]) if len(row)>4 and row[4] else 60
                    })
        
        # Process Rooms
        rooms_sheet = wb["Rooms"] if "Rooms" in wb.sheetnames else None
        rooms_data = []
        if rooms_sheet:
            for row in rooms_sheet.iter_rows(min_row=2, values_only=True):
                if row[0]: # Room Code
                    room_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    room_type = _norm_room_type(row[3]) if len(row) > 3 and row[3] else "classroom"
                    rooms_data.append({
                        "college_id": college_id,
                        "room_code": str(row[0]).strip()[:20],
                        "room_name": room_name[:100],
                        "name": room_name[:100],
                        "capacity": int(row[2]) if len(row)>2 and row[2] else 60,
                        "room_type": room_type,
                        "type": room_type[:10],
                        "available_days": _norm_days(row[4]) if len(row) > 4 and row[4] else []
                    })
        
        def _trim_strings_to_limit(rows, char_limit):
            for row in rows:
                for key, val in list(row.items()):
                    if isinstance(val, str) and len(val) > char_limit:
                        row[key] = val[:char_limit]

        def robust_insert(table_name, data):
            if not data:
                return

            current_data = [row.copy() for row in data]
            while True:
                try:
                    _batch_insert(table_name, current_data, chunk_size=50)
                    break
                except Exception as e:
                    error_msg = str(e)
                    if "PGRST204" in error_msg and "Could not find the" in error_msg:
                        match = re.search(r"Could not find the '([^']+)' column", error_msg)
                        if match:
                            missing_col = match.group(1)
                            for row in current_data:
                                row.pop(missing_col, None)
                            continue
                    # Legacy schema fallback: if DB has tight varchar(N), trim strings and retry.
                    if "value too long for type character varying(" in error_msg:
                        size_match = re.search(r"character varying\((\d+)\)", error_msg)
                        if size_match:
                            _trim_strings_to_limit(current_data, int(size_match.group(1)))
                            continue
                    raise e

        # Keep dataset strictly file-driven: clear the current tenant data after parsing succeeds.
        if replace_existing:
            supabase.table("substitutions").delete().eq("college_id", college_id).execute()
            supabase.table("leave_requests").delete().eq("college_id", college_id).execute()
            supabase.table("timetable_slots").delete().eq("college_id", college_id).execute()
            supabase.table("faculty").delete().eq("college_id", college_id).execute()
            supabase.table("subjects").delete().eq("college_id", college_id).execute()
            supabase.table("rooms").delete().eq("college_id", college_id).execute()
                    
        # Save to database
        robust_insert("faculty", faculty_data)
        robust_insert("subjects", subjects_data)
        robust_insert("rooms", rooms_data)
            
        import json
        import os
        cache_path = os.path.join(os.path.dirname(__file__), "email_cache.json")
        with open(cache_path, "w") as f:
            json.dump(email_cache, f)

        return {
            "message": "Excel data uploaded successfully", 
            "faculty_count": len(faculty_data), 
            "subjects_count": len(subjects_data), 
            "rooms_count": len(rooms_data)
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing Excel file: {str(e)}")
