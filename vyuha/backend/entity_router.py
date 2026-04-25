from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from database import supabase
from dependencies import get_college_id
from auth_system import get_current_user_from_token
from openpyxl import Workbook
from io import BytesIO

router = APIRouter(tags=["Entity Data"])


class FacultyPayload(BaseModel):
    name: str
    employee_id: str = ""
    department: str = ""
    subjects: list = []
    semesters: list = []
    max_classes_per_day: int = 5
    available_days: list = []
    email: str = ""


class SubjectPayload(BaseModel):
    name: str
    code: str = ""
    semester: int
    classes_per_week: int = 2
    room_type_required: str = "classroom"
    duration_minutes: int = 60


class RoomPayload(BaseModel):
    room_code: str
    room_name: str
    capacity: int = 60
    room_type: str = "classroom"
    available_days: list = []


def _require_admin_or_hod(current_user: dict):
    if current_user.get("role") not in ["admin", "hod", "superadmin"]:
        raise HTTPException(status_code=403, detail="Only admin or HOD can modify entities")


def _csv_value(value):
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    return value or ""

@router.get("/faculty")
async def get_faculty(college_id: str = Depends(get_college_id)):
    try:
        res = supabase.table("faculty").select("*").eq("college_id", college_id).execute()
        return res.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching faculty: {str(e)}")


@router.post("/faculty")
async def create_faculty(
    payload: FacultyPayload,
    college_id: str = Depends(get_college_id),
    current_user: dict = Depends(get_current_user_from_token)
):
    _require_admin_or_hod(current_user)
    try:
        result = supabase.table("faculty").insert({
            "college_id": college_id,
            "name": payload.name.strip(),
            "employee_id": payload.employee_id.strip(),
            "department": payload.department.strip(),
            "subjects": payload.subjects or [],
            "semesters": payload.semesters or [],
            "max_classes_per_day": payload.max_classes_per_day,
            "available_days": payload.available_days or [],
            "email": payload.email.strip() if payload.email else None
        }).execute()
        return result.data[0] if result.data else {}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating faculty: {str(e)}")


@router.put("/faculty/{faculty_id}")
async def update_faculty(
    faculty_id: int,
    payload: FacultyPayload,
    college_id: str = Depends(get_college_id),
    current_user: dict = Depends(get_current_user_from_token)
):
    _require_admin_or_hod(current_user)
    try:
        result = supabase.table("faculty").update({
            "name": payload.name.strip(),
            "employee_id": payload.employee_id.strip(),
            "department": payload.department.strip(),
            "subjects": payload.subjects or [],
            "semesters": payload.semesters or [],
            "max_classes_per_day": payload.max_classes_per_day,
            "available_days": payload.available_days or [],
            "email": payload.email.strip() if payload.email else None
        }).eq("id", faculty_id).eq("college_id", college_id).execute()
        if not result.data:
            raise HTTPException(status_code=404, detail="Faculty not found")
        return result.data[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating faculty: {str(e)}")


@router.delete("/faculty/{faculty_id}")
async def delete_faculty(
    faculty_id: int,
    college_id: str = Depends(get_college_id),
    current_user: dict = Depends(get_current_user_from_token)
):
    _require_admin_or_hod(current_user)
    try:
        result = supabase.table("faculty").delete().eq("id", faculty_id).eq("college_id", college_id).execute()
        if not result.data:
            raise HTTPException(status_code=404, detail="Faculty not found")
        return {"message": "Faculty deleted"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting faculty: {str(e)}")

@router.get("/subjects")
async def get_subjects(college_id: str = Depends(get_college_id)):
    try:
        res = supabase.table("subjects").select("*").eq("college_id", college_id).execute()
        return res.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching subjects: {str(e)}")


@router.post("/subjects")
async def create_subject(
    payload: SubjectPayload,
    college_id: str = Depends(get_college_id),
    current_user: dict = Depends(get_current_user_from_token)
):
    _require_admin_or_hod(current_user)
    try:
        result = supabase.table("subjects").insert({
            "college_id": college_id,
            "name": payload.name.strip(),
            "code": payload.code.strip() if payload.code else None,
            "semester": payload.semester,
            "classes_per_week": payload.classes_per_week,
            "room_type_required": payload.room_type_required,
            "duration_minutes": payload.duration_minutes
        }).execute()
        return result.data[0] if result.data else {}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating subject: {str(e)}")


@router.put("/subjects/{subject_id}")
async def update_subject(
    subject_id: int,
    payload: SubjectPayload,
    college_id: str = Depends(get_college_id),
    current_user: dict = Depends(get_current_user_from_token)
):
    _require_admin_or_hod(current_user)
    try:
        result = supabase.table("subjects").update({
            "name": payload.name.strip(),
            "code": payload.code.strip() if payload.code else None,
            "semester": payload.semester,
            "classes_per_week": payload.classes_per_week,
            "room_type_required": payload.room_type_required,
            "duration_minutes": payload.duration_minutes
        }).eq("id", subject_id).eq("college_id", college_id).execute()
        if not result.data:
            raise HTTPException(status_code=404, detail="Subject not found")
        return result.data[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating subject: {str(e)}")


@router.delete("/subjects/{subject_id}")
async def delete_subject(
    subject_id: int,
    college_id: str = Depends(get_college_id),
    current_user: dict = Depends(get_current_user_from_token)
):
    _require_admin_or_hod(current_user)
    try:
        result = supabase.table("subjects").delete().eq("id", subject_id).eq("college_id", college_id).execute()
        if not result.data:
            raise HTTPException(status_code=404, detail="Subject not found")
        return {"message": "Subject deleted"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting subject: {str(e)}")

@router.get("/rooms")
async def get_rooms(college_id: str = Depends(get_college_id)):
    try:
        res = supabase.table("rooms").select("*").eq("college_id", college_id).execute()
        return res.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching rooms: {str(e)}")


@router.get("/export/source-data")
async def export_source_data(
    college_id: str = Depends(get_college_id),
    current_user: dict = Depends(get_current_user_from_token)
):
    _require_admin_or_hod(current_user)

    try:
        faculty_res = supabase.table("faculty").select("*").eq("college_id", college_id).order("id").execute()
        subjects_res = supabase.table("subjects").select("*").eq("college_id", college_id).order("id").execute()
        rooms_res = supabase.table("rooms").select("*").eq("college_id", college_id).order("id").execute()

        wb = Workbook()

        faculty_ws = wb.active
        faculty_ws.title = "Faculty"
        faculty_headers = ["Name", "Employee ID", "Subjects", "Semesters", "Max Classes/Day", "Available Days", "Department", "Email"]
        faculty_ws.append(faculty_headers)
        for row in faculty_res.data:
            faculty_ws.append([
                row.get("name", ""),
                row.get("employee_id", ""),
                _csv_value(row.get("subjects", [])),
                _csv_value(row.get("semesters", [])),
                row.get("max_classes_per_day", 5),
                _csv_value(row.get("available_days", [])),
                row.get("department", ""),
                row.get("email", ""),
            ])

        subjects_ws = wb.create_sheet("Subjects")
        subjects_headers = ["Semester", "Name", "Classes/Week", "Room Type Required", "Duration (min)"]
        subjects_ws.append(subjects_headers)
        for row in subjects_res.data:
            subjects_ws.append([
                row.get("semester", ""),
                row.get("name", ""),
                row.get("classes_per_week", 2),
                row.get("room_type_required", "classroom"),
                row.get("duration_minutes", 60),
            ])

        rooms_ws = wb.create_sheet("Rooms")
        rooms_headers = ["Room Code", "Room Name", "Capacity", "Room Type", "Available Days"]
        rooms_ws.append(rooms_headers)
        for row in rooms_res.data:
            rooms_ws.append([
                row.get("room_code", ""),
                row.get("room_name", row.get("name", "")),
                row.get("capacity", 60),
                row.get("room_type", row.get("type", "classroom")),
                _csv_value(row.get("available_days", [])),
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{college_id}_source_data.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting source data: {str(e)}")


@router.post("/rooms")
async def create_room(
    payload: RoomPayload,
    college_id: str = Depends(get_college_id),
    current_user: dict = Depends(get_current_user_from_token)
):
    _require_admin_or_hod(current_user)
    try:
        result = supabase.table("rooms").insert({
            "college_id": college_id,
            "room_code": payload.room_code.strip(),
            "room_name": payload.room_name.strip(),
            "capacity": payload.capacity,
            "room_type": payload.room_type,
            "available_days": payload.available_days or []
        }).execute()
        return result.data[0] if result.data else {}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating room: {str(e)}")


@router.put("/rooms/{room_id}")
async def update_room(
    room_id: int,
    payload: RoomPayload,
    college_id: str = Depends(get_college_id),
    current_user: dict = Depends(get_current_user_from_token)
):
    _require_admin_or_hod(current_user)
    try:
        result = supabase.table("rooms").update({
            "room_code": payload.room_code.strip(),
            "room_name": payload.room_name.strip(),
            "capacity": payload.capacity,
            "room_type": payload.room_type,
            "available_days": payload.available_days or []
        }).eq("id", room_id).eq("college_id", college_id).execute()
        if not result.data:
            raise HTTPException(status_code=404, detail="Room not found")
        return result.data[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating room: {str(e)}")


@router.delete("/rooms/{room_id}")
async def delete_room(
    room_id: int,
    college_id: str = Depends(get_college_id),
    current_user: dict = Depends(get_current_user_from_token)
):
    _require_admin_or_hod(current_user)
    try:
        result = supabase.table("rooms").delete().eq("id", room_id).eq("college_id", college_id).execute()
        if not result.data:
            raise HTTPException(status_code=404, detail="Room not found")
        return {"message": "Room deleted"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting room: {str(e)}")
